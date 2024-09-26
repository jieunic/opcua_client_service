import os
from openpyxl import load_workbook
from dotenv import load_dotenv

class Excel:
    def import_node_id(file_path, temp_nodes):
        if file_path:
            print(file_path)
            workbook = load_workbook(file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 5:
                    nodeId = row[4]  
                    node = str(nodeId)
                    temp_nodes.append(node)
                else:
                    print(f"Row does not have enough columns: {row}")
        else:
            print('Data file does not exist...')

class Txt:
    def logging():
        print("Done arrival!")

class Env:
    def get(env_path, prefix):
        load_dotenv(env_path)
        val = os.getenv(prefix)
        return val